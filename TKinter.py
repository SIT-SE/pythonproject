# Import
from tkinter import *
# Import allow us to create workbook on our own
from openpyxl.workbook import Workbook
# Import allow us to open excel that has ben created
from openpyxl import load_workbook

# Set up the window (root widget)
root = Tk()

# Open excel sheet
# Create a wb instance to open my excel book
wb = Workbook()
# Now load the excel sheet
# The excel is in GUI directory if not pls specify path
wb = load_workbook('Covid-19 SG.xlsx')

# Create active sheet only one sheet in our case
ws = wb.active

# Create variable for column A
Date = ws['A']
Daily_Confirmed = ws['B']


def get_Date():
    Dlist = ''
    for cell in Date:
        Dlist = f'{Dlist + str(cell.value)}'

    Date_Label.config(text=Dlist)


def get_DC():
    DClist = ''
    for cell in Daily_Confirmed:
        DClist = f'{DClist + str(cell.value)}'

    DC_Label.config(text=DClist)


# Creating buttons to display Date
button_Date = Button(root, text="Date", command=get_Date)
button_Date.pack(pady=20)

# Create a widget: Date
Date_Label = Label(root, text="")
# Pack
Date_Label.pack(pady=20)

# Creating buttons to display Daily Case
button_DC = Button(root, text="Daily_confirmed", command=get_DC)
button_DC.pack(pady=20)

# Create a widget: Label
DC_Label = Label(root, text="")
# Pack
DC_Label.pack(pady=20)

# Create a loop to check for end of program
root.mainloop()
